import requests, json
import pandas as pd
import csv
import re
import sys
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl.utils import get_column_letter

managers = ["Brian","Caoimhín","Niamh","Seán","Violet"]
draft_file="drafted_players_2425.csv"
client_file="client_key.json"
spreadsheet="FPL Draft Stats 2024_25"

def construct_draft_teams():
    draft_teams = {}
    for manager in managers:
        add_draft_team(draft_teams,manager)
    return draft_teams

def add_draft_team(draft_teams,name):
    draft_teams.update({name:{"start_gkp":{},
                              "sub_gkp":{},
                              "def":{},
                              "mid":{},
                              "fwd":{},
                              "subs":{1:{},
                                      2:{},
                                      3:{}}
                              }})


def add_player(draft_teams,manager,position,player,id_num,outfield_sub):
    if(outfield_sub!=0):
        draft_teams[manager]["subs"][outfield_sub]={"name":player,"position":position,"id":id_num,"score":0,"minutes":0}
    else:
        draft_teams[manager][position][player]={}
        draft_teams[manager][position][player]["id"]=id_num
        draft_teams[manager][position][player]["score"]=0
        draft_teams[manager][position][player]["minutes"]=0

def read_player_csv(filename):
    with open(filename, newline='') as f:
        reader = csv.reader(f)
        player_data = list(reader)
    return player_data

def get_player_data():
    base_url = 'https://fantasy.premierleague.com/api/' 
    r = requests.get(base_url+'bootstrap-static/').json() 
    players = pd.json_normalize(r['elements'])
    teams = pd.json_normalize(r['teams'])
    positions = pd.json_normalize(r['element_types'])
    df = pd.merge(
        left=players,
        right=teams,
        left_on='team',
        right_on='id'
    )
    df = df.merge(
        positions,
        left_on='element_type',
        right_on='id'
    )
    df = df.rename(
        columns={'name':'team_name', 'singular_name':'position_name'}
    )
    #print(df[['first_name', 'second_name', 'team_name', 'position_name','id_x']].loc[df['web_name'] == 'Dalot'])
    return df
   
def get_gameweek_history(player_id):
    base_url = 'https://fantasy.premierleague.com/api/' 
    r = requests.get(base_url + 'element-summary/' + str(player_id) + '/').json()
    player_df = pd.json_normalize(r['history'])
    return player_df

def add_gameweek_data(draft_teams,gameweek):
    for manager in managers:
        for position in ["start_gkp","sub_gkp","def","mid","fwd","subs"]:
            for player in draft_teams[manager][position]:
                player_gameweek_history = get_gameweek_history(draft_teams[manager][position][player]["id"])
                draft_teams[manager][position][player]["score"] = player_gameweek_history.loc[player_gameweek_history['round'] == gameweek, 'total_points'].sum()
                draft_teams[manager][position][player]["minutes"] = player_gameweek_history.loc[player_gameweek_history['round'] == gameweek, 'minutes'].sum()

def get_score(team):
    starting_positions = ["def","mid","fwd","start_gkp"]
    positions = ["start_gkp","sub_gkp","def","mid","fwd","subs"]
    min_formation = [3,2,1]
    score = get_initial_score(team)
    formation = get_formation(team)
    num_subs = 0
    for position in starting_positions:
        for player in team[position]:
            if team[position][player]["minutes"] == 0:
                if position == "start_gkp":
                    for player in team["sub_gkp"]:
                        score = score + team["sub_gkp"][player]["score"]
                else:
                    formation[starting_positions.index(position)] = formation[starting_positions.index(position)]-1
                    num_subs = num_subs + 1

    subs_required = [0,0,0]
    for index, position in enumerate(subs_required):
        if min_formation[index] > formation[index]:
            subs_required[index] = min_formation[index] - formation[index]

    while num_subs > 0:
        sub_not_made = True
        if sum(subs_required) > 0:
            for sub in range(1,4):
                if team["subs"][sub]["minutes"] > 0 and subs_required[starting_positions.index(team["subs"][sub]["position"])] > 0:
                    score = score + team["subs"][sub]["score"]
                    num_subs = num_subs - 1
                    team["subs"][sub]["minutes"] = 0
                    subs_required[starting_positions.index(team["subs"][sub]["position"])] = subs_required[starting_positions.index(team["subs"][sub]["position"])] - 1
                    sub_not_made = False
                    break
        else:
            for sub in range(1,4):
                if team["subs"][sub]["minutes"] > 0:
                    score = score + team["subs"][sub]["score"]
                    num_subs = num_subs - 1
                    team["subs"][sub]["minutes"] = 0
                    sub_not_made = False
                    break
        if sub_not_made:
            break

    return int(score)
    

def get_initial_score(team):
    starting_positions = ["start_gkp","def","mid","fwd"]
    initial_score = 0
    for position in starting_positions:
        for player in team[position]:
            initial_score = initial_score + team[position][player]["score"]
    return initial_score

def get_formation(team):
    formation = [len(team["def"]),len(team["mid"]),len(team["fwd"])]
    return formation


def read_file(filename):
    with open(filename, 'r') as fp:
        return(''.join(fp.readlines()))

def extract_data(data):
    players, points = [], []
    score_regex = re.search(" Points</h4><div class=\"EntryEvent__PrimaryValue-ernz96-4 bGEHdY\">(-?\\d+)", data)
    score = int((score_regex.group(1)))
    players_regex = re.findall("([\\w=\\d\\s\\.-]+)</div><div class=\"styles__ElementValue-sc-52mmxp-6 cHYlGH\">(-?\\d*)<",data)
    for match in players_regex:
        if "=" in match[0]: # Handles players with special characters in their names
            unicode_regex = re.findall("=([\\d\\w]{2})=([\\d\\w]{2})",match[0])
            player_name = match[0]
            for unicode_match in unicode_regex:
                player_name = re.sub("=" + unicode_match[0] + "=" + unicode_match[1],bytes.fromhex(unicode_match[0] + unicode_match[1]).decode(),player_name)
            players.append(player_name)
        else:
            players.append(match[0])
        if not match[1]:
            points.append(0)
        else:
            points.append(int(match[1]))
    return players, points, score

def write_squad_to_spreadsheet(client,gameweek,manager,players,points):
    sheet = client.open(spreadsheet).worksheet("Squads")
    manager_index = managers.index(manager)
    row = 2+manager_index*17
    column = 2+(gameweek-1)*2
    cell_string=get_column_letter(column)+str(row)+":"+get_column_letter(column)+str(row+14)
    sheet.update(values=list(map(list, zip(*[players]))),range_name=cell_string)
    cell_string=get_column_letter(column+1)+str(row)+":"+get_column_letter(column+1)+str(row+14)
    sheet.update(values=list(map(list, zip(*[points]))),range_name=cell_string)
    return

def write_scores_to_spreadsheet(client,gameweek,scores,sheetname):
    sheet = client.open(spreadsheet).worksheet(sheetname)
    row = 3
    column = gameweek+1
    cell_string=get_column_letter(column)+str(row)+":"+get_column_letter(column)+str(row+4)
    sheet.update(values=list(map(list, zip(*[scores]))),range_name=cell_string)
    return

def authorise_credentials():
    scope = [
        'https://www.googleapis.com/auth/drive',
        'https://www.googleapis.com/auth/drive.file'
        ]
    creds = ServiceAccountCredentials.from_json_keyfile_name(client_file,scope)
    client = gspread.authorize(creds)
    return client

def process_standard(gameweek,client):
    standard_scores = []
    for manager in managers:
        filename = manager + ".mhtml"
        print("\t\tParsing " + manager + ".mhtml")
        data = read_file(filename)
        players, points, score = extract_data(data)
        print("\t\tDone")
        standard_scores.append(score)
        print("\t\tWriting " + manager + "'s squad to spreadsheet")
        write_squad_to_spreadsheet(client,gameweek,manager,players,points)
        print("\t\tDone")
    print("\t\tWriting standard scores to spreadsheet")
    write_scores_to_spreadsheet(client,gameweek,standard_scores,"Scores")
    print("\t\tDone")
    return

def process_saf(gameweek,client):
    saf_scores = []
    draft_teams = construct_draft_teams()
    print("\t\tReading Set-And-Forget teams from file")
    player_data = read_player_csv(draft_file)
    print("\t\tDone")
    print("\t\tAdding players to data structure")
    for player in player_data:
        add_player(draft_teams,player[0],player[1],player[2],int(player[3]),int(player[4]))
    print("\t\tDone")
    print("\t\tGetting players' data for game week " + str(gameweek))
    add_gameweek_data(draft_teams,gameweek)
    print("\t\tDone")
    for manager in managers:
        saf_scores.append(get_score(draft_teams[manager]))
    print("\t\tWriting scores to spreadsheet")
    write_scores_to_spreadsheet(client,gameweek,saf_scores,"SetAndForgetScores")
    print("\t\tDone")
    return

if __name__ == "__main__":
    gameweek = int(sys.argv[1])

    print("\tAuthorising credentials")
    client = authorise_credentials()
    print("\tDone")
    print("\tProcessing standard scores and squads")
    process_standard(gameweek,client)
    print("\tDone")
    print("\tProcessing Set-And-Forget scores")
    process_saf(gameweek,client)
    print("\tDone")
