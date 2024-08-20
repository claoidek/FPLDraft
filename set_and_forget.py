import requests, json
import pandas as pd
import csv
import sys
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl.utils import get_column_letter

managers = ["Brian","Caoimhín","Niamh","Seán","Violet"]

def construct_draft_teams(managers):
    draft_teams = {}
    for manager in managers:
        add_draft_team(draft_teams,manager)
    return draft_teams

def add_draft_team(draft_teams,name):
    draft_teams.update({name:{"start_gkp":{},
                              "sub_gkp":{},
                              "def":{},
                              "mid":{},
                              "fwd":{},
                              "subs":{1:{},
                                      2:{},
                                      3:{}}
                              }})


def add_player(draft_teams,manager,position,player,id_num,outfield_sub):
    if(outfield_sub!=0):
        draft_teams[manager]["subs"][outfield_sub]={"name":player,"position":position,"id":id_num,"score":0,"minutes":0}
    else:
        draft_teams[manager][position][player]={}
        draft_teams[manager][position][player]["id"]=id_num
        draft_teams[manager][position][player]["score"]=0
        draft_teams[manager][position][player]["minutes"]=0

def read_player_csv(filename):
    with open(filename, newline='') as f:
        reader = csv.reader(f)
        player_data = list(reader)
    return player_data

def get_player_data():
    base_url = 'https://fantasy.premierleague.com/api/' 
    r = requests.get(base_url+'bootstrap-static/').json() 
    players = pd.json_normalize(r['elements'])
    teams = pd.json_normalize(r['teams'])
    positions = pd.json_normalize(r['element_types'])
    df = pd.merge(
        left=players,
        right=teams,
        left_on='team',
        right_on='id'
    )
    df = df.merge(
        positions,
        left_on='element_type',
        right_on='id'
    )
    df = df.rename(
        columns={'name':'team_name', 'singular_name':'position_name'}
    )
    #print(df[['first_name', 'second_name', 'team_name', 'position_name','id_x']].loc[df['web_name'] == 'Dalot'])
    return df
   
def get_gameweek_history(player_id):
    base_url = 'https://fantasy.premierleague.com/api/' 
    r = requests.get(base_url + 'element-summary/' + str(player_id) + '/').json()
    player_df = pd.json_normalize(r['history'])
    return player_df

def add_gameweek_data(draft_teams,gameweek,managers):
    for manager in managers:
        for position in ["start_gkp","sub_gkp","def","mid","fwd","subs"]:
            for player in draft_teams[manager][position]:
                player_gameweek_history = get_gameweek_history(draft_teams[manager][position][player]["id"])
                draft_teams[manager][position][player]["score"] = player_gameweek_history.loc[player_gameweek_history['round'] == gameweek, 'total_points'].sum()
                draft_teams[manager][position][player]["minutes"] = player_gameweek_history.loc[player_gameweek_history['round'] == gameweek, 'minutes'].sum()

def get_score(team):
    starting_positions = ["def","mid","fwd","start_gkp"]
    positions = ["start_gkp","sub_gkp","def","mid","fwd","subs"]
    min_formation = [3,2,1]
    score = get_initial_score(team)
    formation = get_formation(team)
    num_subs = 0
    for position in starting_positions:
        for player in team[position]:
            if team[position][player]["minutes"] == 0:
                if position == "start_gkp":
                    for player in team["sub_gkp"]:
                        score = score + team["sub_gkp"][player]["score"]
                else:
                    formation[starting_positions.index(position)] = formation[starting_positions.index(position)]-1
                    num_subs = num_subs + 1

    subs_required = [0,0,0]
    for index, position in enumerate(subs_required):
        if min_formation[index] > formation[index]:
            subs_required[index] = min_formation[index] - formation[index]

    while num_subs > 0:
        sub_not_made = True
        if sum(subs_required) > 0:
            for sub in range(1,4):
                if team["subs"][sub]["minutes"] > 0 and subs_required[starting_positions.index(team["subs"][sub]["position"])] > 0:
                    score = score + team["subs"][sub]["score"]
                    num_subs = num_subs - 1
                    team["subs"][sub]["minutes"] = 0
                    subs_required[starting_positions.index(team["subs"][sub]["position"])] = subs_required[starting_positions.index(team["subs"][sub]["position"])] - 1
                    sub_not_made = False
                    break
        else:
            for sub in range(1,4):
                if team["subs"][sub]["minutes"] > 0:
                    score = score + team["subs"][sub]["score"]
                    num_subs = num_subs - 1
                    team["subs"][sub]["minutes"] = 0
                    sub_not_made = False
                    break
        if sub_not_made:
            break

    return int(score)
    

def get_initial_score(team):
    starting_positions = ["start_gkp","def","mid","fwd"]
    initial_score = 0
    for position in starting_positions:
        for player in team[position]:
            initial_score = initial_score + team[position][player]["score"]
    return initial_score

def get_formation(team):
    formation = [len(team["def"]),len(team["mid"]),len(team["fwd"])]
    return formation

def write_scores_to_spreadsheet(client,gameweek,scores):
    sheet = client.open('FPL Draft Stats 2024_25').worksheet("SetAndForgetScores")
    row = 3
    column = gameweek+1
    cell_string=get_column_letter(column)+str(row)+":"+get_column_letter(column)+str(row+4)
    sheet.update(values=list(map(list, zip(*[scores]))),range_name=cell_string)

def authorise_credentials():
    scope = [
        'https://www.googleapis.com/auth/drive',
        'https://www.googleapis.com/auth/drive.file'
        ]
    file_name = 'client_key.json'
    creds = ServiceAccountCredentials.from_json_keyfile_name(file_name,scope)
    client = gspread.authorize(creds)
    return client

if __name__ == "__main__":
    gameweek = int(sys.argv[1])
    scores = []
    print("\tAuthorising credentials")
    client = authorise_credentials()
    print("\tDone")
    draft_teams = construct_draft_teams(managers)
    print("\tReading Set-And-Forget teams from file")
    player_data = read_player_csv("drafted_players_2425.csv")
    print("\tDone")
    print("\tAdding players to data structure")
    for player in player_data:
        add_player(draft_teams,player[0],player[1],player[2],int(player[3]),int(player[4]))
    print("\tDone")
    print("\tGetting players' data for game week " + str(gameweek))
    add_gameweek_data(draft_teams,gameweek,managers)
    print("\tDone")
    for manager in managers:
        scores.append(get_score(draft_teams[manager]))
    print("\tWriting scores to spreadsheet")
    write_scores_to_spreadsheet(client,gameweek,scores)
    print("\tDone")
