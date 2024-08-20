import re
import sys
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl.utils import get_column_letter

managers = ["Brian","Caoimhín","Niamh","Seán","Violet"]

def read_file(filename):
    with open(filename, 'r') as fp:
        important_lines = fp.readlines()
        no_equals = [sub[: -2] for sub in important_lines] # Removes the last character of every line to get rid of inconvenient equals signs
        return(''.join(no_equals))

def extract_data(data):
    players, points = [], []
    score_regex = re.search(" Points</h4><div class=\"EntryEvent__PrimaryValue-ernz96-4 bGEHdY\">(-?\\d+)", data)
    score = int((score_regex.group(1)))
    players_regex = re.findall("([\\w=\\d\\s\\.-]+)</div><div class=\"styles__ElementValue-sc-52mmxp-6 cHYlGH\">(-?\\d*)<",data)
    for match in players_regex:
        if "=" in match[0]: # Handles players with special characters in their names
            unicode_regex = re.findall("=([\\d\\w]{2})=([\\d\\w]{2})",match[0])
            player_name = match[0]
            for unicode_match in unicode_regex:
                player_name = re.sub("=" + unicode_match[0] + "=" + unicode_match[1],bytes.fromhex(unicode_match[0] + unicode_match[1]).decode(),player_name)
            players.append(player_name)
        else:
            players.append(match[0])
        if not match[1]:
            points.append(0)
        else:
            points.append(int(match[1]))
    return players, points, score

def write_squad_to_spreadsheet(client,gameweek,manager,players,points):
    sheet = client.open('FPL Draft Stats 2024_25').worksheet("Squads")
    manager_index = managers.index(manager)
    row = 2+manager_index*17
    column = 2+(gameweek-1)*2
    cell_string=get_column_letter(column)+str(row)+":"+get_column_letter(column)+str(row+14)
    sheet.update(values=list(map(list, zip(*[players]))),range_name=cell_string)
    cell_string=get_column_letter(column+1)+str(row)+":"+get_column_letter(column+1)+str(row+14)
    sheet.update(values=list(map(list, zip(*[points]))),range_name=cell_string)
    return

def write_scores_to_spreadsheet(client,gameweek,scores):
    sheet = client.open('FPL Draft Stats 2024_25').worksheet("Scores")
    row = 3
    column = gameweek+1
    cell_string=get_column_letter(column)+str(row)+":"+get_column_letter(column)+str(row+4)
    sheet.update(values=list(map(list, zip(*[scores]))),range_name=cell_string)

def authorise_credentials():
    scope = [
        'https://www.googleapis.com/auth/drive',
        'https://www.googleapis.com/auth/drive.file'
        ]
    file_name = 'client_key.json'
    creds = ServiceAccountCredentials.from_json_keyfile_name(file_name,scope)
    client = gspread.authorize(creds)
    return client

if __name__ == "__main__":
    gameweek = int(sys.argv[1])
    scores = []
    print("\tAuthorising credentials")
    client = authorise_credentials()
    print("\tDone")
    for manager in managers:
        filename = manager + ".mhtml"
        print("\tParsing " + manager + "'s mhtml file")
        data = read_file(filename)
        players, points, score = extract_data(data)
        print("\tDone")
        scores.append(score)
        print("\tWriting " + manager + "'s squad to spreadsheet")
        write_squad_to_spreadsheet(client,gameweek,manager,players,points)
        print("\tDone")
    print("\tWriting scores to spreadsheet")
    write_scores_to_spreadsheet(client,gameweek,scores)
    print("\tDone")
